def columnTidier (workbook):  #Puesto a mano, luego vemos de automatizarlo.
    for sheet in workbook:
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 50
        sheet.column_dimensions['F'].width = 17
        sheet.column_dimensions['H'].width = 3
        sheet.column_dimensions['I'].width = 10
        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['K'].width = 10
        sheet.column_dimensions['L'].width = 15


def write2excel (promos_database, pasadas_totales, pasadas_por_horarios, destination_filename):
    '''
    IN: 
    lista_pasadas = list of Pasadas objects
    promos_database = dict
    
    filename = (str) excel filename (include .xlsx extension)
    OUT: Writes Excel file, with separate sheet for each feed
    '''
    encabezadoIBMS = [
    'MediaMI', 'Descripcion', 'Duracion', 'Pasadas totales', '00:00-06:00',
    '06:00-12:00', '12:00-18:00', '18:00-24', 'Descripcion larga', 'Comentarios'
    ]

    from openpyxl import Workbook
    from constants import BLOQUES_HORARIOS

    wb = Workbook() #creo libro nuevo . 
    #la plani por default se llama 'Sheet'. Al final de todo, hay que borrarla, porque esta vacía

    for pasadas_keys, pasadas_values in pasadas_totales.items():
        feed = pasadas_keys[0]
        media_MI = pasadas_keys[1]
        descripcion = list(promos_database.get(media_MI))[0]
        duracion = list(promos_database.get(media_MI))[1]
        pasadas_totales = pasadas_values
        pasadas00_06 = pasadas_por_horarios.get((feed, media_MI, BLOQUES_HORARIOS[0]), 0)
        pasadas06_12 = pasadas_por_horarios.get((feed, media_MI, BLOQUES_HORARIOS[1]), 0)
        pasadas12_18 = pasadas_por_horarios.get((feed, media_MI, BLOQUES_HORARIOS[2]), 0)
        pasadas18_24 = pasadas_por_horarios.get((feed, media_MI, BLOQUES_HORARIOS[3]), 0)
        descripcion_larga = list(promos_database.get(media_MI))[2]

        promoRow = [media_MI, descripcion, duracion, pasadas_totales, pasadas00_06, pasadas06_12, pasadas12_18, pasadas18_24, descripcion_larga]

        print (encabezadoIBMS)
        print (promoRow)

        if feed not in wb.sheetnames: #si no hay una hoja con el nombre de esa señal
            wb.create_sheet(title = feed)  #creo hoja nueva con el nombre de esa señal
            ws1 = wb[feed] #asigno a la variable ws1 el nombre de esa plani
            ws1.append(encabezadoIBMS) #agrego el encabezado al comienzo de la lista
            ws1.append(promoRow)  #la agrego como fila a la planilla ws1
        else:
            ws1 = wb[feed] #asigno a la variable ws1 el nombre de esa plani
            ws1.append(promoRow)  #la agrego como fila a la planilla ws1

    # Al final de todo, borrar la Hoja 'Sheet' que esta vacía
    wb.remove(wb['Sheet'])

    # Ajustar (manualmente) el ancho de las columnas de todas las hojas
    columnTidier(wb)
    #creo el archivo ("w" especifica que si no existe, lo cree)
    f = open(destination_filename, "w+")
    #Salvo el libro
    wb.save(filename = destination_filename)
